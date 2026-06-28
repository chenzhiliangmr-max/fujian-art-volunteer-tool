import json
import re
import subprocess
from collections import defaultdict
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
PLAN_XLSX = ROOT / "outputs" / "fujian-art-2026-plan-review-draft.xlsx"
OUT_XLSX = ROOT / "outputs" / "fujian-art-2026-plan-vs-2025-admission-review.xlsx"
OUT_CSV = ROOT / "outputs" / "fujian-art-2026-plan-vs-2025-admission-review.csv"


def norm_text(value):
    text = "" if value is None else str(value)
    text = text.strip()
    replacements = {
        chr(0xFF08): "(",
        chr(0xFF09): ")",
        chr(0x3008): "(",
        chr(0x3009): ")",
        chr(0x3014): "(",
        chr(0x3015): ")",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"\s+", "", text)
    return text


def norm_school(value):
    text = norm_text(value)
    text = text.replace(chr(0x798F) + chr(0x5EFA) + chr(0x7701), chr(0x798F) + chr(0x5EFA))
    typo = "".join(chr(x) for x in [0x4E0A,0x4FA8,0x4E2D,0x4FA8,0x804C,0x4E1A,0x6280,0x672F,0x5927,0x5B66])
    fixed = "".join(chr(x) for x in [0x4E0A,0x6D77,0x4E2D,0x4FA8,0x804C,0x4E1A,0x6280,0x672F,0x5927,0x5B66])
    text = text.replace(typo, fixed)
    return text


def norm_major(value):
    text = norm_text(value)
    text = text.replace(chr(0x4E13) + chr(0x4E1A), "")
    return text


def major_base(value):
    text = norm_major(value)
    text = text.split("(")[0]
    fragment = "".join(chr(x) for x in [0x5883,0x8BBE,0x8BA1,0x4EA7,0x54C1,0x8BBE,0x8BA1,0x5DE5,0x827A,0x7F8E,0x672F])
    if fragment in text:
        return "".join(chr(x) for x in [0x8BBE,0x8BA1,0x5B66,0x7C7B])
    return text


def parse_int(value):
    if value is None:
        return None
    match = re.search(r"\d+", str(value))
    return int(match.group(0)) if match else None


def read_2026_plan():
    wb = openpyxl.load_workbook(PLAN_XLSX, data_only=True)
    ws = wb["PLAN_DRAFT"]
    rows = []
    for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        subject_cn = str(values[0] or "").strip()
        subject = str(values[1] or "").strip()
        if not subject:
            subject = {chr(0x5386)+chr(0x53F2): "history", chr(0x7269)+chr(0x7406): "physics"}.get(subject_cn, subject)
        item = {
            "excel_row": excel_row,
            "subject": subject,
            "subject_cn": subject_cn,
            "page": values[2],
            "school_code": str(values[4] or "").strip(),
            "school": str(values[5] or "").strip(),
            "province": str(values[6] or "").strip(),
            "level": str(values[7] or "").strip(),
            "site": str(values[8] or "").strip(),
            "group": str(values[9] or "").strip(),
            "major_code": str(values[10] or "").strip(),
            "major": str(values[11] or "").strip(),
            "plan_count": parse_int(values[12]),
            "duration": str(values[13] or "").strip(),
            "tuition": str(values[14] or "").strip(),
            "note": str(values[15] or "").strip(),
        }
        if item["school"] and item["major"]:
            rows.append(item)
    return rows


def read_2025_admission():
    script = """
const data = require('./shared/fujian-art-data.js');
const rows = data.colleges.filter((row) => Number(row.year) === 2025);
process.stdout.write(JSON.stringify(rows));
"""
    result = subprocess.run(
        ["node", "-e", script],
        cwd=ROOT,
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    raw_rows = json.loads(result.stdout)
    rows = []
    for row in raw_rows:
        info = str(row.get("info") or "")
        marker = chr(0xFF0C) + chr(0x5F55) + chr(0x53D6)
        major = info.split(marker)[0].strip() if marker in info else info.strip()
        admit_count = None
        match = re.search(r"录取\s*(\d+)", info)
        if match:
            admit_count = int(match.group(1))
        rows.append(
            {
                "subject": str(row.get("subject") or "").strip(),
                "school": str(row.get("school") or "").strip(),
                "province": str(row.get("province") or "").strip(),
                "level": str(row.get("level") or "").strip(),
                "major": major,
                "admit_count": admit_count,
                "min": row.get("min"),
                "max": row.get("max"),
                "avg": row.get("avg"),
                "info": info,
                "status": str(row.get("status") or "").strip(),
            }
        )
    return rows


def make_key(row, school_field="school", major_field="major"):
    return (
        str(row.get("subject") or "").strip(),
        norm_school(row.get(school_field)),
        norm_major(row.get(major_field)),
    )


def make_base_key(row, school_field="school", major_field="major"):
    return (
        str(row.get("subject") or "").strip(),
        norm_school(row.get(school_field)),
        major_base(row.get(major_field)),
    )


def index_rows(rows, key_func):
    grouped = defaultdict(list)
    for row in rows:
        grouped[key_func(row)].append(row)
    return grouped


def best_match(plan, admission_by_exact, admission_by_base):
    exact = admission_by_exact.get(make_key(plan))
    if exact:
        return "exact", exact
    base = admission_by_base.get(make_base_key(plan))
    if base:
        return "base_major", base
    return "none", []


def add_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    for col in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in col) + 2, 44)
        ws.column_dimensions[col[0].column_letter].width = width
    return ws


def main():
    plan_2026 = read_2026_plan()
    admission_2025 = read_2025_admission()

    plan_exact = index_rows(plan_2026, make_key)
    admission_exact = index_rows(admission_2025, make_key)
    plan_base = index_rows(plan_2026, make_base_key)
    admission_base = index_rows(admission_2025, make_base_key)

    matched_plan_keys = set()
    matched_admission_keys = set()
    match_rows = []
    new_major_rows = []

    for plan in plan_2026:
        mode, matches = best_match(plan, admission_exact, admission_base)
        key = make_key(plan)
        if mode == "none":
            new_major_rows.append(plan)
            continue
        matched_plan_keys.add(key)
        for adm in matches:
            matched_admission_keys.add(make_key(adm))
            diff = None
            if plan["plan_count"] is not None and adm["admit_count"] is not None:
                diff = plan["plan_count"] - adm["admit_count"]
            match_rows.append((plan, adm, mode, diff))

    cancelled_major_rows = []
    for adm in admission_2025:
        if make_key(adm) in matched_admission_keys:
            continue
        if make_key(adm) in plan_exact:
            continue
        if make_base_key(adm) in plan_base:
            continue
        cancelled_major_rows.append(adm)

    plan_schools = {(r["subject"], norm_school(r["school"])) for r in plan_2026}
    admission_schools = {(r["subject"], norm_school(r["school"])) for r in admission_2025}
    new_school_keys = sorted(plan_schools - admission_schools)
    cancelled_school_keys = sorted(admission_schools - plan_schools)

    new_school_rows = []
    for subject, school in new_school_keys:
        items = [r for r in plan_2026 if r["subject"] == subject and norm_school(r["school"]) == school]
        new_school_rows.append([subject, items[0]["subject_cn"], items[0]["school"], len(items), sum(i["plan_count"] or 0 for i in items)])

    cancelled_school_rows = []
    for subject, school in cancelled_school_keys:
        items = [r for r in admission_2025 if r["subject"] == subject and norm_school(r["school"]) == school]
        cancelled_school_rows.append([subject, items[0]["school"], len(items), sum(i["admit_count"] or 0 for i in items)])

    count_change_rows = []
    for plan, adm, mode, diff in match_rows:
        if diff is None or diff == 0:
            continue
        count_change_rows.append(
            [
                plan["subject"],
                plan["school"],
                plan["major"],
                plan["plan_count"],
                adm["admit_count"],
                diff,
                "扩招" if diff > 0 else "缩招",
                mode,
                adm["min"],
                adm["avg"],
            ]
        )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    add_sheet(
        wb,
        "SUMMARY",
        ["项目", "数量"],
        [
            ["2026计划行数", len(plan_2026)],
            ["2025录取行数", len(admission_2025)],
            ["2026新增学校(按科类)", len(new_school_rows)],
            ["2025有但2026取消学校(按科类)", len(cancelled_school_rows)],
            ["2026新增专业/方向", len(new_major_rows)],
            ["2025有但2026取消专业/方向", len(cancelled_major_rows)],
            ["同专业人数变化", len(count_change_rows)],
            ["匹配成功关系", len(match_rows)],
        ],
    )
    add_sheet(wb, "NEW_SCHOOLS_2026", ["subject", "科类", "院校名称", "2026专业数", "2026计划合计"], new_school_rows)
    add_sheet(wb, "CANCELLED_SCHOOLS_2026", ["subject", "院校名称", "2025专业数", "2025录取合计"], cancelled_school_rows)
    add_sheet(
        wb,
        "NEW_MAJORS_2026",
        ["subject", "科类", "页码", "院校代码", "院校名称", "专业组", "专业代码", "专业名称", "2026计划人数", "学制", "学费", "备注"],
        [
            [r["subject"], r["subject_cn"], r["page"], r["school_code"], r["school"], r["group"], r["major_code"], r["major"], r["plan_count"], r["duration"], r["tuition"], r["note"]]
            for r in new_major_rows
        ],
    )
    add_sheet(
        wb,
        "CANCELLED_MAJORS_2026",
        ["subject", "院校名称", "专业名称", "2025录取人数", "2025最低分", "2025平均分", "2025信息"],
        [[r["subject"], r["school"], r["major"], r["admit_count"], r["min"], r["avg"], r["info"]] for r in cancelled_major_rows],
    )
    add_sheet(
        wb,
        "COUNT_CHANGES",
        ["subject", "院校名称", "专业名称", "2026计划人数", "2025录取人数", "变化", "类型", "匹配方式", "2025最低分", "2025平均分"],
        sorted(count_change_rows, key=lambda x: (x[0], x[1], x[2])),
    )
    add_sheet(
        wb,
        "MATCH_DETAIL",
        [
            "subject",
            "院校名称",
            "2026专业名称",
            "2026计划人数",
            "2025专业名称",
            "2025录取人数",
            "人数变化",
            "匹配方式",
            "2025最低分",
            "2025平均分",
        ],
        [
            [p["subject"], p["school"], p["major"], p["plan_count"], a["major"], a["admit_count"], diff, mode, a["min"], a["avg"]]
            for p, a, mode, diff in match_rows
        ],
    )
    wb.save(OUT_XLSX)

    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        import csv

        writer = csv.writer(f)
        writer.writerow(["sheet", "subject", "school", "major", "detail"])
        for r in new_major_rows:
            writer.writerow(["NEW_MAJORS_2026", r["subject"], r["school"], r["major"], f"plan={r['plan_count']}"])
        for r in cancelled_major_rows:
            writer.writerow(["CANCELLED_MAJORS_2026", r["subject"], r["school"], r["major"], f"admit={r['admit_count']}"])

    print(json.dumps({
        "plan_2026": len(plan_2026),
        "admission_2025": len(admission_2025),
        "new_schools": len(new_school_rows),
        "cancelled_schools": len(cancelled_school_rows),
        "new_majors": len(new_major_rows),
        "cancelled_majors": len(cancelled_major_rows),
        "count_changes": len(count_change_rows),
        "out_xlsx": str(OUT_XLSX),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
