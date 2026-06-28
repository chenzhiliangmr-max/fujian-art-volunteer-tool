#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build review files for 2026 Fujian art admission plan PDFs.

The source PDFs are scanned image PDFs. OCR is done separately by Windows OCR
into C:\tmp\fujian-art-2026-plan-ocr-lines-xy.json. This script reconstructs
page rows by OCR coordinates and creates human-review files. It deliberately
does not import data into the product.
"""

from __future__ import annotations

import csv
import json
import re
import shutil
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OCR_PATH = Path(r"C:\tmp\fujian-art-2026-plan-ocr-lines-xy.json")
OUT_DIR = ROOT / "outputs"
OUT_XLSX = OUT_DIR / "fujian-art-2026-plan-review-draft.xlsx"
OUT_PLAN_CSV = OUT_DIR / "fujian-art-2026-plan-standard-draft.csv"
OUT_MATCH_CSV = OUT_DIR / "fujian-art-2026-plan-vs-2025-match-draft.csv"
OUT_OCR_JSON = OUT_DIR / "fujian-art-2026-plan-ocr-lines-xy.json"


DIGIT_MAP = str.maketrans(
    {
        "O": "0",
        "o": "0",
        "〇": "0",
        "０": "0",
        "一": "1",
        "l": "1",
        "I": "1",
        "佣": "0",
        "巳": "已",
        "１": "1",
        "２": "2",
        "３": "3",
        "４": "4",
        "５": "5",
        "６": "6",
        "７": "7",
        "８": "8",
        "９": "9",
    }
)


def tight(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).replace("\u3000", " "))


def norm_digits(value: object) -> str:
    return tight(value).translate(DIGIT_MAP)


def clean_text(value: object) -> str:
    text = tight(value).translate(DIGIT_MAP)
    return (
        text.replace("：", ":")
        .replace("（", "(")
        .replace("）", ")")
        .replace("，", ",")
        .replace("•", ".")
        .replace("·", ".")
        .replace("／", "/")
        .replace("十", "+")
    )


def norm_key(value: object) -> str:
    text = clean_text(value)
    text = re.sub(r"[()（）\s·.•:：,，。；;、/\\\-—_]+", "", text)
    return text.replace("中外合作办学", "中外合作").replace("闽台合作办学", "闽台合作")


def has_chinese(value: object) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", str(value or "")))


def first_num(value: object) -> str:
    match = re.search(r"\d+", norm_digits(value))
    return match.group(0) if match else ""


def parse_tuition_remark(value: object) -> tuple[str, str]:
    text = clean_text(value)
    if not text:
        return "", ""
    match = re.match(r"^(\d[\d\sO0佣０-９]*)", text)
    if match:
        return norm_digits(match.group(1)), text[match.end() :].strip(" ,，;；")
    return "", text


def group_page_rows(items: list[dict], tolerance: int = 18) -> list[dict]:
    items = sorted(items, key=lambda row: (float(row["y"]), float(row["x"])))
    groups: list[dict] = []
    for item in items:
        y = float(item["y"])
        if not groups or abs(groups[-1]["y"] - y) > tolerance:
            groups.append({"y": y, "items": [item]})
        else:
            groups[-1]["items"].append(item)
            groups[-1]["y"] = (
                groups[-1]["y"] * (len(groups[-1]["items"]) - 1) + y
            ) / len(groups[-1]["items"])
    return [
        {
            "y": round(group["y"], 1),
            "items": sorted(group["items"], key=lambda row: float(row["x"])),
        }
        for group in groups
    ]


def col_text(items: list[dict], left: int, right: int) -> str:
    return "".join(
        clean_text(item.get("text", ""))
        for item in items
        if left <= float(item["x"]) < right
    )


def row_text(items: list[dict]) -> str:
    return " | ".join(
        f"[{int(float(item['x']))}]{item.get('text', '')}"
        for item in sorted(items, key=lambda row: float(row["x"]))
    )


def load_2025_data() -> list[dict]:
    data_path = ROOT / "shared" / "fujian-art-data.js"
    text = data_path.read_text(encoding="utf-8")
    return json.loads(text[text.index("[") : text.rindex("];") + 1])


def extract_plan(raw: list[dict]) -> tuple[list[dict], list[dict]]:
    by_page: dict[tuple[str, int], list[dict]] = defaultdict(list)
    for row in raw:
        by_page[(row["subject"], int(row["page"]))].append(row)

    page_rows_out: list[dict] = []
    plan_records: list[dict] = []
    school_words = ("大学", "学院", "学校", "美院", "艺院")
    ignore_words = (
        "招生章程",
        "院校",
        "代号",
        "专业",
        "名称",
        "计划",
        "人数",
        "收费",
        "标准",
        "备注",
        "福建省普通高校",
    )

    for subject, subject_name in (("history", "历史"), ("physics", "物理")):
        current_school = ""
        current_school_code = ""
        current_group = ""
        current_url = ""
        current_location = ""
        current_nature = ""
        current_major: dict | None = None
        active = False
        pages = sorted(page for subj, page in by_page if subj == subject)

        for page in pages:
            rows = group_page_rows(by_page[(subject, page)])
            if page > 1:
                active = True
            for seq, row in enumerate(rows, 1):
                items = row["items"]
                rebuilt = row_text(items)
                all_clean = clean_text("".join(item.get("text", "") for item in items))
                left_col = col_text(items, 0, 380)
                major_col = col_text(items, 380, 1120)
                school_col = col_text(items, 220, 390)
                year_col = col_text(items, 1120, 1260)
                plan_col = col_text(items, 1260, 1370)
                tuition_col = col_text(items, 1370, 1535)
                remark_col = col_text(items, 1535, 2400)

                page_rows_out.append(
                    {
                        "科类": subject_name,
                        "subject": subject,
                        "页码": page,
                        "页内行序": seq,
                        "y": row["y"],
                        "左侧/院校列": left_col,
                        "专业列": major_col,
                        "学制列": year_col,
                        "计划人数列": plan_col,
                        "收费/地区列": tuition_col,
                        "备注列": remark_col,
                        "重建原文": rebuilt,
                    }
                )

                if not active:
                    if "美术与设计类" in all_clean:
                        active = True
                        current_school = ""
                        current_school_code = ""
                        current_group = ""
                        current_url = ""
                        current_location = ""
                        current_nature = ""
                        current_major = None
                    continue

                if not all_clean or any(
                    word in all_clean
                    for word in ("院校专业", "专业名称", "计划人数", "收费标准")
                ):
                    continue
                if "美术与设计类" in all_clean:
                    current_major = None
                    continue
                if "招生章程网址" in all_clean:
                    url = all_clean.split("招生章程网址", 1)[-1].strip(":：")
                    if url:
                        current_url = url
                    continue

                first_x = float(items[0]["x"]) if items else 9999
                maybe_school = clean_text(school_col or items[0].get("text", ""))
                if (
                    first_x < 390
                    and has_chinese(maybe_school)
                    and any(word in maybe_school for word in school_words)
                    and not any(word in maybe_school for word in ignore_words)
                    and not re.match(r"^\d{3,4}", norm_digits(maybe_school))
                ):
                    current_school = maybe_school
                    current_major = None
                    locs = re.findall(r"\(([^)]*)\)", clean_text(tuition_col + remark_col))
                    current_location = locs[0] if locs else ""
                    current_nature = locs[1] if len(locs) > 1 else ""
                    continue

                if "专业组" in all_clean:
                    current_major = None
                    match = re.search(
                        r"([0-9O〇佣lI\s]{2,5})专业组[:：]?([0-9O〇佣lI\s]{1,5})?",
                        all_clean,
                    )
                    if match:
                        school_code = re.sub(r"\D", "", norm_digits(match.group(1)))
                        group_code = re.sub(r"\D", "", norm_digits(match.group(2) or ""))
                        if school_code:
                            current_school_code = school_code
                        if group_code:
                            current_group = group_code
                    continue

                major_code = ""
                major_name = ""
                match = re.match(r"^([0-9O〇佣lI\s]{2,4})(.+)$", (major_col or all_clean).strip())
                if match:
                    major_code = re.sub(r"\D", "", norm_digits(match.group(1)))
                    major_name = clean_text(match.group(2))
                else:
                    major_items = [
                        item for item in items if 380 <= float(item["x"]) < 1120
                    ]
                    if len(major_items) >= 2:
                        major_code = re.sub(
                            r"\D", "", norm_digits(major_items[0].get("text", ""))
                        )
                        major_name = clean_text(
                            "".join(item.get("text", "") for item in major_items[1:])
                        )

                if major_code and has_chinese(major_name):
                    plan_count = first_num(plan_col)
                    tuition, tuition_remark = parse_tuition_remark(tuition_col)
                    remark = (
                        tuition_remark
                        + (
                            ("；" if tuition_remark and remark_col else "") + remark_col
                            if remark_col
                            else ""
                        )
                    ).strip("；")
                    status: list[str] = []
                    if not current_school:
                        status.append("缺少院校名称")
                    if not current_school_code:
                        status.append("缺少院校代码")
                    if not current_group:
                        status.append("缺少专业组")
                    if current_school_code and len(current_school_code) != 4:
                        status.append("\u9662\u6821\u4ee3\u7801\u7591\u4f3cOCR\u6f0f\u4f4d")
                    if not plan_count:
                        status.append("缺少计划人数")
                    if len(major_code) != 3:
                        status.append("专业代码疑似OCR异常")
                    if tuition.isdigit() and int(tuition) < 5000:
                        status.append("\u5b66\u8d39\u7591\u4f3cOCR\u6f0f\u96f6")
                    if current_school and page > 1 and seq < 5:
                        status.append("页首记录疑似承接上一页")

                    rec = {
                        "科类": subject_name,
                        "subject": subject,
                        "页码": page,
                        "页内行序": seq,
                        "院校代码": current_school_code,
                        "院校名称": current_school,
                        "地区": current_location,
                        "办学性质": current_nature,
                        "招生章程网址OCR": current_url,
                        "专业组": current_group,
                        "专业代码": major_code.zfill(3)
                        if major_code.isdigit() and len(major_code) <= 3
                        else major_code,
                        "专业名称": major_name,
                        "计划人数": plan_count,
                        "学制": clean_text(year_col),
                        "学费": tuition,
                        "备注": remark,
                        "跨页说明": "承接上一页" if page > 1 and seq < 5 else "",
                        "解析状态": "待复核：" + "；".join(status)
                        if status
                        else "初步解析",
                        "原始行": rebuilt,
                    }
                    plan_records.append(rec)
                    current_major = rec
                    continue

                if current_major and all_clean:
                    min_x = min(float(item["x"]) for item in items)
                    if min_x >= 1500 or (
                        min_x >= 500 and not any(word in all_clean for word in ("专业组", "招生章程网址"))
                    ):
                        if min_x < 1150 and not plan_col and not tuition_col and not remark_col:
                            current_major["专业名称"] += all_clean
                        else:
                            current_major["备注"] = (
                                current_major.get("备注", "") + all_clean
                            ).strip()
                            if current_major["解析状态"] == "初步解析":
                                current_major["解析状态"] = "初步解析（含续行备注）"

    return page_rows_out, plan_records


def build_match_rows(plan_records: list[dict]) -> list[dict]:
    existing = load_2025_data()
    by_exact: dict[tuple[str, str, str], list[dict]] = defaultdict(list)
    by_school: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in existing:
        if int(row.get("year", 0)) != 2025:
            continue
        subject = row.get("subject", "")
        school = row.get("school", "")
        major = re.split("[，,]", row.get("info", ""))[0]
        item = {
            "2025院校": school,
            "2025专业": major,
            "2025录取人数信息": row.get("info", ""),
            "2025最低分": row.get("min", ""),
            "2025平均分": row.get("avg", ""),
            "2025最高分": row.get("max", ""),
            "2025状态": row.get("status", ""),
            "2025标签": ",".join(row.get("tags", [])),
        }
        by_exact[(subject, norm_key(school), norm_key(major))].append(item)
        by_school[(subject, norm_key(school))].append(item)

    match_rows: list[dict] = []
    for rec in plan_records:
        matches = by_exact.get(
            (rec["subject"], norm_key(rec["院校名称"]), norm_key(rec["专业名称"])),
            [],
        )
        if matches:
            status = "2025同校同专业匹配"
            candidates = matches
        else:
            candidates = by_school.get((rec["subject"], norm_key(rec["院校名称"])), [])
            status = "仅匹配到2025同校，专业需确认" if candidates else "2025未匹配到同校同专业"

        if candidates:
            for match in candidates[:5]:
                match_rows.append({**rec, **match, "匹配状态": status})
        else:
            match_rows.append(
                {
                    **rec,
                    "2025院校": "",
                    "2025专业": "",
                    "2025录取人数信息": "",
                    "2025最低分": "",
                    "2025平均分": "",
                    "2025最高分": "",
                    "2025状态": "",
                    "2025标签": "",
                    "匹配状态": status,
                }
            )
    return match_rows


def write_csv(path: Path, rows: list[dict]) -> None:
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def add_sheet(wb: Workbook, title: str, rows: list[dict]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["无数据"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8F2F1")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "科类": 8,
        "subject": 9,
        "页码": 8,
        "页内行序": 8,
        "y": 8,
        "院校名称": 24,
        "专业名称": 32,
        "备注": 60,
        "原始行": 90,
        "重建原文": 100,
        "招生章程网址OCR": 45,
        "解析状态": 32,
        "跨页说明": 14,
        "匹配状态": 24,
        "2025录取人数信息": 34,
        "2025状态": 26,
    }
    for idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(header, 14)


def write_workbook(page_rows: list[dict], plan_records: list[dict], match_rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    for row in [
        ["文件用途", "2026福建美术与设计类招生计划 OCR 整理草稿，供人工核查；确认前不接入系统。"],
        ["来源PDF", "2026历史组组美术与设计类福建招生计划.pdf；2026物理组美术与设计类福建招生计划.pdf"],
        ["OCR方式", "PDF为扫描图片，无文本层；使用 Windows 内置中文 OCR，并按页面坐标重建表格。"],
        ["页码处理", "所有 OCR 行和计划记录均保留科类、页码、页内行序。页首专业行会标记“承接上一页”。"],
        ["统计-历史组OCR行", sum(1 for row in page_rows if row["subject"] == "history")],
        ["统计-物理组OCR行", sum(1 for row in page_rows if row["subject"] == "physics")],
        ["统计-计划记录", len(plan_records)],
        ["统计-匹配记录", len(match_rows)],
    ]:
        ws.append(row)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 100
    add_sheet(wb, "OCR_ROWS", page_rows)
    add_sheet(wb, "PLAN_DRAFT", plan_records)
    add_sheet(wb, "MATCH_2025", match_rows)
    review_rows = [
        {
            key: row.get(key, "")
            for key in [
                "科类",
                "页码",
                "页内行序",
                "院校代码",
                "院校名称",
                "专业组",
                "专业代码",
                "专业名称",
                "计划人数",
                "学制",
                "学费",
                "解析状态",
                "跨页说明",
                "原始行",
            ]
        }
        for row in plan_records
        if "待复核" in row.get("解析状态", "")
        or "续行" in row.get("解析状态", "")
        or row.get("跨页说明")
    ]
    add_sheet(wb, "REVIEW_FIRST", review_rows)
    wb.save(OUT_XLSX)


def main() -> None:
    OUT_DIR.mkdir(exist_ok=True)
    raw = json.loads(OCR_PATH.read_text(encoding="utf-8-sig"))
    shutil.copyfile(OCR_PATH, OUT_OCR_JSON)
    page_rows, plan_records = extract_plan(raw)
    match_rows = build_match_rows(plan_records)
    write_csv(OUT_PLAN_CSV, plan_records)
    write_csv(OUT_MATCH_CSV, match_rows)
    write_workbook(page_rows, plan_records, match_rows)
    print(f"plan_records={len(plan_records)}")
    print(f"match_rows={len(match_rows)}")
    print(OUT_XLSX)
    print(OUT_PLAN_CSV)
    print(OUT_MATCH_CSV)


if __name__ == "__main__":
    main()
